import io
from openpyxl import load_workbook
from app.pii_engine import sanitize_batch


def parse_xlsx(data: bytes, mode: str = "redact") -> tuple[bytes, int]:
    wb = load_workbook(io.BytesIO(data))

    # Collect all string cells for batch processing
    all_texts = []
    cell_refs = []  # store actual cell objects
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    all_texts.append(cell.value)
                    cell_refs.append(cell)

    if not all_texts:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), 0

    # Single batch sanitize
    cleaned_texts, total = sanitize_batch(all_texts, mode)

    for i, cell in enumerate(cell_refs):
        cell.value = cleaned_texts[i]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), total
